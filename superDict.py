# -*- coding: utf-8 -*-
import json

import openpyxl
from openpyxl import load_workbook

#此方法用于更新词库
def importDict(mode):
    #xlsxPath = 'D:\Mirai\YirisVoiceGUI\PythonPlugins\Config\可爱系二次元bot词库1.5万词V1.1.xlsx'
    if mode==1:
        xlsxPath = 'Config\词库.xlsx'
    else:
        xlsxPath = 'Config\完全匹配.xlsx'
    # 第一步打开工作簿
    wb = openpyxl.load_workbook(xlsxPath)
    # 第二步选取表单
    sheet = wb.active
    # 按行获取数据转换成列表
    rows_data = list(sheet.rows)
    # 获取表单的表头信息(第一行)，也就是列表的第一个元素
    titles = [title.value for title in rows_data.pop(0)]
    print(titles)


    all_row_dict = []

    #当你需要向已有词库导入时取消注释
    '''fileaa = open('Config\\superDict.txt', 'r')
    js1 = fileaa.read()
    newDict = json.loads(js1)
    print('已读取现存字典')'''

    #新建词库，当你需要新建词库时取消注释
    newDict={}

    # 遍历出除了第一行的其他行
    for a_row in rows_data:
        the_row_data = [cell.value for cell in a_row]
        # 将表头和该条数据内容，打包成一个字典
        row_dict = dict(zip(titles, the_row_data))
        #print(row_dict)
        all_row_dict.append(row_dict)
    for i in all_row_dict:

        key=i.get('key')
        value=i.get('value')

        if (key in newDict):
            replyValue=newDict.get(key)
            if value in replyValue:
                print('已存在该回复，不添加')
            else:
                replyValue.append(value)
                print('已有关键字，追加')
            #print(replyValue)
        # 没有关键字则创建
        else:
            newDict[key] = [value,]
        #print('key:'+key+' '+'value:'+value)
    print(newDict)
    js = json.dumps(newDict)
    if mode==1:
        file = open('Config\superDict.txt', 'w')
    else:
        file = open('Config\Dict.txt', 'w')
    file.write(js)
    file.close()

def clearSheet(a):
    i=0
    if a==1:
        filename = 'Config/词库.xlsx'
    else:
        filename= 'Config/完全匹配.xlsx'
    while i<4:

        wb = load_workbook(filename)
        ws = wb.active
        ws.delete_cols(1)  # 删除第 1 列数据
        wb.save(filename)
        i+=1
        print('clear')
def outPutDic(a):
    if a==1:
        filename = 'Config/完全匹配.xlsx'
        file = open('Config\\dict.txt', 'r')
        js = file.read()
        dict = json.loads(js)
        Keys = dict.keys()
        print('已读取字典')
    else:


        filename = 'Config/词库.xlsx'
        file = open('Config\\superDict.txt', 'r')
        jss = file.read()
        dict = json.loads(jss)
        Keys = dict.keys()
        print('已读取模糊匹配字典')



    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    sheet.append(['key', 'value'])  # 插入一行数据
    wb.save(filename)  # 保存,传入原文件则在

    for d in Keys:
        key=d
        values=dict.get(d)

        for value in values:
            wb = openpyxl.load_workbook(filename)
            print(str(key)+str(value))
            sheet = wb.active


            sheet.append([key, value])  # 插入一行数据
            wb.save(filename)  # 保存,传入原文件则在




if __name__ == '__main__':
    importDict(1)#从excel导入词库
    #importDict(2)
    #clearSheet(1)
    #clearSheet(2)
    #outPutDic(1)#导出到excel
    #outPutDic(2)
